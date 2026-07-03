"""UMBS bronze schema management — loader + generators in one place.

This module owns three concerns:

1. **Loading** the canonical bronze schema from `schemas.yaml`.
   `BRONZE_SCHEMAS[gse][folder][col] = pl.DataType` is exposed at module
   load via an `lru_cache`'d YAML reader. Consumed by `import_bronze.py`
   to pin polars dtypes at scan time, preventing the silent-null inference
   bugs we hit pre-2026-05-25.

2. **Building** `schemas.yaml` from Fannie's UMBS Disclosure Guide
   (`schemas/umbs/mbsglossary.xlsx`). Single Security Initiative (2019)
   means the same Fannie spec covers both Fannie and Freddie UMBS files.
   Hand-curated naming aliases live in `spec_aliases.yaml`; the
   sheet→folder mapping is the `SHEET_TO_FOLDERS` constant below
   (structural metadata, not data).

3. **Discovering** dtypes empirically by full-scanning every raw zip
   with everything-as-String and inferring narrowest types that hold
   across all months. Slower (~80 min) but data-truth-derived. Used as
   a verification cross-check against the spec build, and as a seed of
   defaults for columns the spec doesn't cover (Filler fields, multifamily
   extras, newer ESG fields, file-header typos).

Entry points:
    python -m mortgage_data_manager.umbs.schemas              # build from spec
    python -m mortgage_data_manager.umbs.schemas build        # explicit
    python -m mortgage_data_manager.umbs.schemas discover     # full-scan
"""

from __future__ import annotations

import argparse
import tempfile
import time
import zipfile
from functools import lru_cache
from pathlib import Path
from typing import Any

import polars as pl
import yaml
from openpyxl import load_workbook

from mortgage_data_manager.core.logging import configure_logging, get_logger
from mortgage_data_manager.umbs.config import UMBSConfig

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Paths and constants
# ---------------------------------------------------------------------------

_HERE = Path(__file__).parent
_SCHEMAS_YAML = _HERE / "schemas.yaml"
_ALIASES_YAML = _HERE / "spec_aliases.yaml"
_DISCOVERED_YAML = _HERE / "schemas_discovered.yaml"
_GLOSSARY_PATH = UMBSConfig.PROJECT_DIR / "schemas" / "umbs" / "mbsglossary.xlsx"

# Polars dtype <-> YAML name. We only persist the three dtypes that appear
# in UMBS disclosure feeds — String, Int64, Float64. Dates and decimals are
# stored as Int64/Float64 (lossless for the integer-formatted dates Fannie
# uses, e.g. ``"122049"`` for Dec 2049).
_DTYPE_NAMES: dict[pl.DataType, str] = {
    pl.String: "String",
    pl.Int64: "Int64",
    pl.Float64: "Float64",
}
_NAMES_TO_DTYPE: dict[str, pl.DataType] = {v: k for k, v in _DTYPE_NAMES.items()}

# Type-promotion lattice for discovery merge across files: highest rank wins.
# "Empty" means a column was 100% null/empty in the scanned file and
# contributes no information; merged into String at the end (safe default).
_DISCOVERY_RANK = {"Empty": 0, "Int64": 1, "Float64": 2, "String": 3}

# Which glossary sheet governs which bronze folder. Structural metadata that
# ties spec sheets to disclosure-feed folders; not data that should live in
# YAML. Update when adding new folders to PIPE_DELIMITED_WITH_HEADERS /
# COMMA_DELIMITED_FOLDERS in import_bronze.py.
SHEET_TO_FOLDERS: dict[str, list[tuple[str, str]]] = {
    "7-SC Loan Level ": [
        ("FNMA", "FNM_ILLD"), ("FNMA", "FNM_MLLD"), ("FNMA", "FNM_RILLD"),
        ("FHLMC", "FRE_ILLD"), ("FHLMC", "FRE_RILLD"),
        ("FHLMC", "FU"), ("FHLMC", "AC"), ("FHLMC", "AU"),
    ],
    "8-SC Security Level": [
        ("FNMA", "FNM_IS"), ("FNMA", "FNM_MF"), ("FNMA", "FNM_RIS"),
        ("FHLMC", "FRE_IS"), ("FHLMC", "FRE_RIS"),
        ("FHLMC", "FD"), ("FHLMC", "AR"), ("FHLMC", "XF"),
        # MI/MW are FHLMC multifamily — share the security-level base schema
        # plus a few MF-specific extras handled via discovery fallback.
        ("FHLMC", "MI"), ("FHLMC", "MW"),
    ],
    "15-SC Daily Prepayment Report ": [
        ("FNMA", "FNM_DPR_FCTR"), ("FHLMC", "FRE_DPR_Fctr"),
    ],
    "16-GinnieMae Backed Megas ": [
        ("FNMA", "FNM_GN_MEGA"),
    ],
    "21-MC REMIC Factor": [
        ("FNMA", "FNM_REMIC"),
        # FHLMC's REMIC tranche family share the same 17-col schema.
        ("FHLMC", "MT"), ("FHLMC", "GT"), ("FHLMC", "MD"),
        ("FHLMC", "TB"), ("FHLMC", "TQ"), ("FHLMC", "TV"),
        ("FHLMC", "FT"), ("FHLMC", "WJ"),
    ],
    "22-MC SMBS Factor": [
        ("FNMA", "FNM_SMBS"),
        ("FHLMC", "SF"), ("FHLMC", "SX"),
    ],
    "23-MC Component Factor": [
        ("FNMA", "FNM_REMIC_COMPONENT"),
        ("FHLMC", "WL"), ("FHLMC", "CG"),
    ],
    "24- MC REMIC Shortfall ": [
        ("FNMA", "FNM_REMIC_SHORT"),
    ],
    "25-MC ESF Security File": [
        ("FNMA", "FNM_ESF_MS"),
    ],
    "28-MC Floater Reset Indices": [
        ("FNMA", "FNM_RESET_D"), ("FNMA", "FNM_RESET_ND"),
        # FHLMC's index files share the same 4-col layout.
        ("FHLMC", "ID"), ("FHLMC", "IR"), ("FHLMC", "IS"),
        ("FHLMC", "DR"), ("FHLMC", "MR"), ("FHLMC", "OP"),
    ],
    "29-MC Floater Reset Rates": [
        ("FNMA", "FNM_RESET_D_RATES"), ("FNMA", "FNM_RESET_ND_RATES"),
        ("FHLMC", "RD"), ("FHLMC", "DC"),
    ],
}


# ---------------------------------------------------------------------------
# Loader — what import_bronze actually consumes
# ---------------------------------------------------------------------------

@lru_cache(maxsize=1)
def _read_yaml(path_str: str) -> dict[str, Any]:
    """Load and cache a YAML file by absolute path string."""
    path = Path(path_str)
    if not path.exists():
        return {}
    with open(path) as f:
        return yaml.safe_load(f) or {}


def _load_bronze_schemas() -> dict[str, dict[str, dict[str, pl.DataType]]]:
    """Load `schemas.yaml` and convert dtype names → polars dtype objects."""
    raw = _read_yaml(str(_SCHEMAS_YAML))
    out: dict[str, dict[str, dict[str, pl.DataType]]] = {}
    for gse, folders in raw.items():
        out[gse] = {}
        for folder, cols in folders.items():
            out[gse][folder] = {
                col: _NAMES_TO_DTYPE[dt] for col, dt in cols.items()
            }
    return out


def _load_aliases() -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    """Load `spec_aliases.yaml` → (per-sheet alias maps, legacy renames)."""
    raw = _read_yaml(str(_ALIASES_YAML))
    return (
        raw.get("sheet_aliases", {}) or {},
        raw.get("legacy_inherits_from_spec", {}) or {},
    )


# Module-level export: same shape and name as the old auto-generated module,
# so `from mortgage_data_manager.umbs.schemas import BRONZE_SCHEMAS` continues
# to work for every existing consumer.
BRONZE_SCHEMAS: dict[str, dict[str, dict[str, pl.DataType]]] = _load_bronze_schemas()


# ---------------------------------------------------------------------------
# Glossary parsing
# ---------------------------------------------------------------------------

def _spec_to_polars(data_type: Any, fmt: Any) -> pl.DataType | None:
    """Map a glossary (Data Type, Format) pair to a polars dtype.

    Rules:
        - ``String``  → ``pl.String``
        - ``Date``    → ``pl.Int64`` (dates are stored as integer-formatted
                        strings like ``"122049"`` for Dec 2049; Int64 is the
                        lossless storage choice).
        - ``Numeric`` with a fractional format (``"9.2"``, ``"2.3"``, etc.)
                      → ``pl.Float64``
        - ``Numeric`` with no format or an integer-only format → ``pl.Int64``
        - blank / whitespace / unknown → ``None`` (caller falls back)
    """
    if data_type is None or str(data_type).strip() == "":
        return None
    dt = str(data_type).strip()
    if dt == "String":
        return pl.String
    if dt == "Date":
        return pl.Int64
    if dt == "Numeric":
        if fmt is None:
            return pl.Int64
        if isinstance(fmt, (int, float)):
            return pl.Float64 if fmt != int(fmt) else pl.Int64
        if "." in str(fmt):
            return pl.Float64
        return pl.Int64
    return None


def _parse_glossary_sheet(
    sheet_name: str, glossary_path: Path
) -> dict[str, pl.DataType | None]:
    """Return ``{spec_name: polars_dtype}`` for one glossary sheet."""
    wb = load_workbook(glossary_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"Sheet {sheet_name!r} not found in {glossary_path}. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        i for i, r in enumerate(rows[:10]) if r and "Attribute Name" in r
    )
    h = rows[header_idx]
    name_col = h.index("Attribute Name")
    dtype_col = h.index("Data Type") if "Data Type" in h else None
    fmt_col = h.index("Format") if "Format" in h else None

    out: dict[str, pl.DataType | None] = {}
    for r in rows[header_idx + 1:]:
        if not r or not r[name_col]:
            continue
        out[str(r[name_col]).strip()] = _spec_to_polars(
            r[dtype_col] if dtype_col is not None else None,
            r[fmt_col] if fmt_col is not None else None,
        )
    return out


# ---------------------------------------------------------------------------
# YAML writing
# ---------------------------------------------------------------------------

def _write_yaml(
    schemas: dict[str, dict[str, dict[str, str]]],
    output_path: Path,
    header_lines: list[str],
) -> None:
    """Write a nested ``{gse: {folder: {col: dtype_name}}}`` dict to YAML."""
    body = yaml.safe_dump(
        schemas, sort_keys=False, default_flow_style=False, width=200
    )
    header = "\n".join(f"# {line}" if line else "#" for line in header_lines)
    output_path.write_text(header + "\n\n" + body)


# ---------------------------------------------------------------------------
# build_from_spec — canonical path: glossary xlsx → schemas.yaml
# ---------------------------------------------------------------------------

def build_from_spec(
    glossary_path: Path = _GLOSSARY_PATH,
    output_path: Path = _SCHEMAS_YAML,
    discovered_seed: dict[str, dict[str, dict[str, str]]] | None = None,
) -> dict[str, list[str]]:
    """Build ``schemas.yaml`` from the Fannie UMBS glossary spec.

    For each ``(gse, folder, col)`` triple seen in the discovery seed, pick
    the spec dtype if the column matches a glossary entry (directly, via an
    alias in ``spec_aliases.yaml``, or via the legacy-rename inheritance
    map). Fall back to the discovery-derived dtype for spec-uncovered
    columns (Filler fields, multifamily extras, ESG additions, header typos).

    Args:
        glossary_path: Path to ``mbsglossary.xlsx``.
        output_path: Where to write ``schemas.yaml``.
        discovered_seed: Optional discovery output (``{gse: {folder: {col:
            dtype_name}}}``) supplying the column-presence universe and
            fallback dtypes. Loads ``schemas_discovered.yaml`` if omitted.

    Returns:
        ``{f"{gse}/{folder}": [unmatched_col_names]}`` for monitoring.
    """
    sheet_aliases, legacy_inherits = _load_aliases()
    # Seed precedence: explicit arg → discovery YAML → current schemas.yaml.
    # The third fallback lets ``build`` re-run after a glossary or alias edit
    # without forcing an 80-minute discovery scan; the existing schemas.yaml
    # already encodes the column universe per folder.
    if discovered_seed is not None:
        seed = discovered_seed
    elif _DISCOVERED_YAML.exists():
        seed = _read_yaml(str(_DISCOVERED_YAML))
    elif _SCHEMAS_YAML.exists():
        logger.info(
            f"No discovery seed at {_DISCOVERED_YAML.name}; "
            f"using existing {_SCHEMAS_YAML.name} as the column-universe seed."
        )
        seed = _read_yaml(str(_SCHEMAS_YAML))
    else:
        raise RuntimeError(
            f"No seed available — neither {_DISCOVERED_YAML} nor "
            f"{_SCHEMAS_YAML} exists, and no seed was provided. Run "
            f"`python -m mortgage_data_manager.umbs.schemas discover` to "
            f"bootstrap the column universe from raw files."
        )

    out_schemas: dict[str, dict[str, dict[str, str]]] = {}
    warnings: dict[str, list[str]] = {}

    for sheet_name, folders in SHEET_TO_FOLDERS.items():
        spec = _parse_glossary_sheet(sheet_name, glossary_path)
        aliases = sheet_aliases.get(sheet_name, {})
        file_to_spec = {file_h: spec_h for spec_h, file_h in aliases.items()}

        for gse, folder in folders:
            seed_cols = seed.get(gse, {}).get(folder)
            if not seed_cols:
                logger.warning(
                    f"  {gse}/{folder}: no discovery seed; skipping"
                )
                continue

            folder_schema: dict[str, str] = {}
            for col, seed_name in seed_cols.items():
                dtype: pl.DataType | None = None
                if col in spec:
                    dtype = spec[col]
                elif col in file_to_spec:
                    dtype = spec.get(file_to_spec[col])
                elif col in legacy_inherits:
                    dtype = spec.get(legacy_inherits[col])

                if dtype is not None:
                    folder_schema[col] = _DTYPE_NAMES[dtype]
                else:
                    folder_schema[col] = seed_name  # discovery fallback
                    if not col.startswith("Filler"):
                        warnings.setdefault(
                            f"{gse}/{folder}", []
                        ).append(col)

            out_schemas.setdefault(gse, {})[folder] = folder_schema

    # Sort GSE / folder keys so YAML output is deterministic across rebuilds.
    out_schemas = {
        gse: dict(sorted(out_schemas[gse].items()))
        for gse in sorted(out_schemas)
    }

    _write_yaml(
        out_schemas, output_path,
        header_lines=[
            "UMBS bronze layer column schemas.",
            "",
            "AUTO-GENERATED — DO NOT EDIT BY HAND. To regenerate, run:",
            "    python -m mortgage_data_manager.umbs.schemas build",
            "",
            "Source of truth: Fannie Mae UMBS Disclosure Guide",
            "(schemas/umbs/mbsglossary.xlsx). The Single Security Initiative",
            "(2019) harmonized Fannie and Freddie UMBS disclosure formats, so",
            "this single spec governs both GSEs' files.",
            "",
            "Hand-curated overrides (cosmetic glossary-vs-file-header drift,",
            "legacy column renames) live in spec_aliases.yaml.",
            "",
            "Shape: gse → folder → column → polars dtype name",
            '("String" | "Int64" | "Float64"). Loaded by',
            "umbs.schemas._load_bronze_schemas() with lru_cache and consumed",
            "by import_bronze.py to pin polars dtypes at scan time.",
        ],
    )
    logger.info(f"Wrote {output_path}")
    return warnings


# ---------------------------------------------------------------------------
# discover — verification cross-check (full-scan, ~80 min)
# ---------------------------------------------------------------------------

def _infer_one_file(zip_path: Path, separator: str) -> dict[str, str]:
    """Per-column inferred dtype name for one raw zip.

    Streams the file with every column forced to String, then checks per
    column whether all non-empty values cast cleanly to Int64 (with a
    leading-zero check to preserve string-ID semantics) or Float64. Returns
    the narrowest dtype name that holds.
    """
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        candidates = [n for n in names if n.lower().endswith((".txt", ".csv"))]
        data_name = candidates[0] if candidates else names[0]

        with tempfile.TemporaryDirectory() as td:
            zf.extract(data_name, td)
            extracted = Path(td) / data_name

            cols = (
                pl.scan_csv(extracted, separator=separator, has_header=True,
                            n_rows=1)
                .collect_schema().names()
            )

            string_schema = {c: pl.String for c in cols}
            lf = pl.scan_csv(
                extracted, separator=separator, has_header=True,
                schema=string_schema, truncate_ragged_lines=True,
                ignore_errors=False,
            )

            exprs = []
            for i, c in enumerate(cols):
                trimmed = pl.col(c).str.strip_chars()
                is_value = trimmed.is_not_null() & (trimmed != "")
                exprs.append(is_value.cast(pl.Int64).sum().alias(f"_n_{i}"))
                exprs.append(
                    (is_value & trimmed.cast(pl.Int64, strict=False).is_not_null())
                    .cast(pl.Int64).sum().alias(f"_int_{i}")
                )
                exprs.append(
                    (is_value & trimmed.cast(pl.Float64, strict=False).is_not_null())
                    .cast(pl.Int64).sum().alias(f"_flt_{i}")
                )
                # Leading-zero check: preserves IDs like "00123" as String
                # even though they'd parse as Int64.
                lz = (
                    is_value & (trimmed.str.len_chars() > 1)
                    & trimmed.str.starts_with("0")
                    & ~trimmed.str.contains(r"\.", literal=False)
                )
                exprs.append(lz.cast(pl.Int64).sum().alias(f"_lz_{i}"))

            row = lf.select(exprs).collect().row(0)

            result: dict[str, str] = {}
            for i, c in enumerate(cols):
                n, n_int, n_flt, n_lz = (
                    row[i*4], row[i*4+1], row[i*4+2], row[i*4+3]
                )
                if n == 0:
                    result[c] = "Empty"
                elif n_int == n and n_lz == 0:
                    result[c] = "Int64"
                elif n_flt == n:
                    result[c] = "Float64"
                else:
                    result[c] = "String"
            return result


def _merge_discovery(per_file: list[dict[str, str]]) -> dict[str, str]:
    """Take the widest dtype across files for each column."""
    merged: dict[str, str] = {}
    all_cols: set[str] = set()
    for s in per_file:
        all_cols |= s.keys()
    for c in all_cols:
        best = max(
            (s.get(c, "Empty") for s in per_file),
            key=lambda t: _DISCOVERY_RANK[t],
        )
        # All-empty columns default to String (safe if values appear later).
        merged[c] = "String" if best == "Empty" else best
    return merged


def discover(
    raw_dir: Path | None = None,
    output_path: Path = _DISCOVERED_YAML,
) -> dict[str, dict[str, dict[str, str]]]:
    """Full-scan dtype discovery across every UMBS raw zip.

    Writes ``schemas_discovered.yaml`` (the verification cross-check + seed
    for spec-uncovered columns). ~80 min on a Mac SSD; most time spent on
    the two billion-row giants (FNM_MLLD and FU).

    Returns the discovered schema as a nested dict for in-process use.
    """
    # Local import to avoid circular import at module load.
    from mortgage_data_manager.umbs.import_bronze import (
        PIPE_DELIMITED_WITH_HEADERS,
        _separator_for,
    )

    raw_dir = raw_dir or UMBSConfig.UMBS_RAW_DIR
    schemas: dict[str, dict[str, dict[str, str]]] = {}

    folder_universe: dict[str, list[str]] = {}
    for gse, folders in PIPE_DELIMITED_WITH_HEADERS.items():
        folder_universe.setdefault(gse, []).extend(folders)

    for gse in sorted(folder_universe):
        schemas.setdefault(gse, {})
        for folder in sorted(folder_universe[gse]):
            sep = _separator_for(gse, folder)
            folder_dir = raw_dir / gse / folder
            if not folder_dir.exists():
                logger.warning(f"  {gse}/{folder}: raw dir missing, skipping")
                continue
            zips = sorted(folder_dir.glob("*.zip"))
            if not zips:
                logger.warning(f"  {gse}/{folder}: no zips, skipping")
                continue

            t0 = time.time()
            per_file = []
            for zp in zips:
                try:
                    per_file.append(_infer_one_file(zp, sep))
                except Exception as exc:
                    logger.error(f"  {gse}/{folder}/{zp.name}: failed — {exc}")
            if not per_file:
                continue
            schemas[gse][folder] = _merge_discovery(per_file)

            elapsed = time.time() - t0
            counts = {
                t: sum(1 for v in schemas[gse][folder].values() if v == t)
                for t in ["Int64", "Float64", "String"]
            }
            logger.info(
                f"  {gse}/{folder}: {len(zips)} files, "
                f"{len(schemas[gse][folder])} cols  "
                f"[Int64={counts['Int64']} Float64={counts['Float64']} "
                f"String={counts['String']}]  ({elapsed:.1f}s)"
            )

    _write_yaml(
        schemas, output_path,
        header_lines=[
            "UMBS bronze layer column schemas — discovery output.",
            "",
            "AUTO-GENERATED — DO NOT EDIT BY HAND. To regenerate, run:",
            "    python -m mortgage_data_manager.umbs.schemas discover",
            "",
            "Full-scan dtype inference across every raw zip with all columns",
            "forced to String. Slow (~80 min) but data-truth-derived.",
            "",
            "This file is used by build_from_spec as a seed (column universe",
            "+ fallback dtypes for spec-uncovered columns) and as a cross-",
            "check artifact diffable against schemas.yaml.",
        ],
    )
    logger.info(f"Wrote {output_path}")
    return schemas


# ---------------------------------------------------------------------------
# CLI: python -m mortgage_data_manager.umbs.schemas [build|discover]
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> None:
    """Entry point for ``python -m mortgage_data_manager.umbs.schemas``."""
    parser = argparse.ArgumentParser(
        description="UMBS bronze-schema management",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "action", nargs="?", default="build", choices=["build", "discover"],
        help="build (default): parse glossary spec → schemas.yaml; "
             "discover: full-scan inference → schemas_discovered.yaml",
    )
    args = parser.parse_args(argv)
    configure_logging(level="INFO")
    if args.action == "build":
        warnings = build_from_spec()
        if warnings:
            logger.info("")
            logger.info("Discovery-fallback columns (no spec authority):")
            for folder_key, cols in warnings.items():
                logger.info(f"  {folder_key}: {len(cols)} cols")
                for c in cols[:3]:
                    logger.info(f"    - {c}")
                if len(cols) > 3:
                    logger.info(f"    ... and {len(cols) - 3} more")
    elif args.action == "discover":
        discover()


if __name__ == "__main__":
    main()
